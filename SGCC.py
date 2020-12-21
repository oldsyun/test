import requests
import json
from datetime import datetime
import time
import os
import openpyxl
import sys
import signal
import win32api,win32con
 

headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36 Edg/87.0.664.60",
"Content-Type": "application/json",
"Referer": "https://ecp.sgcc.com.cn/ecp2.0/portal/",
}
codelist=[]
flag=0


def getNoticeBid(lstid):
    getNoticeBidurl='https://ecp.sgcc.com.cn/ecp2.0/ecpwcmcore//index/getNoticeBid'
    res=json.loads(requests.post(url=getNoticeBidurl,data=json.dumps(lstid), headers=headers).text)
    PurType=res["resultValue"]["notice"]["PUR_TYPE_NAME"]
    bidlist=[]
    bidlist.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    bidlist.append(res["resultValue"]["notice"]["PURPRJ_NAME"])
    bidlist.append(res["resultValue"]["notice"]["PURPRJ_CODE"])
    bidlist.append(res["resultValue"]["notice"]["BIDBOOK_BUY_END_TIME"])
    bidlist.append(res["resultValue"]["notice"]["BIDBOOK_SELL_BEGIN_TIME"])
    zipname=str(res["resultValue"]["notice"]["ONLINE_BID_NOTICE_ID"])+".zip"
    bidlist.append('=HYPERLINK("Download\\'+zipname+'","'+zipname+'")')
    if PurType=="物资":
        print("新增物资采购 \r"+datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        writexls(bidlist,"物资采购")
    else:
        print("新增服务采购 \r"+datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        writexls(bidlist,"服务采购")
    getDownloadurl(lstid)
        
 
def getChangeBid(lstid): 
    getChangeBidurl='https://ecp.sgcc.com.cn/ecp2.0/ecpwcmcore//index/getChangeBid'
    res=json.loads(requests.post(url=getChangeBidurl,data=json.dumps(lstid), headers=headers).text)
    #print(res)
    return res

def writexls(Bid,sheetname):
    wb = openpyxl.load_workbook(os.getcwd()+'\\Res\\采购信息.xlsx')
    ws = wb[sheetname]
    if sheetname=="物资采购" or sheetname=="服务采购":
        ws['A1'] = '获取时间'
        ws.column_dimensions['A'].width = 23
        ws['B1'] = '采购项目名称'
        ws.column_dimensions['B'].width = 72
        ws['C1'] = '采购项目编号'
        ws.column_dimensions['C'].width = 18
        ws['D1'] = '采购文件获取截止时间'
        ws.column_dimensions['D'].width = 23
        ws['E1'] = '开启应答文件时间'
        ws.column_dimensions['E'].width = 23
        ws['F1'] = '公告文件'
        ws.column_dimensions['F'].width = 23  
        ws.style = "Hyperlink"
    ws.append(Bid)
    wb.save(os.getcwd()+'\\Res\\采购信息.xlsx')

def getDownloadurl(lstid):
    downloadurl='https://ecp.sgcc.com.cn/ecp2.0/ecpwcmcore//index/downLoadBid?noticeId='+str(lstid)+'&noticeDetId='
    cmd = 'Res\\wget --content-disposition "%s" -O "%s"' %(downloadurl,"Download\\"+str(lstid)+".zip")
    os.system(cmd)

def purchase():
    url = 'https://ecp.sgcc.com.cn/ecp2.0/ecpwcmcore//index/noteList'
    #Pre-qualification='{"index":1,"size":20,"firstPageMenuId":"2018032700290425","purOrgStatus":"","purOrgCode":"","purType":"","orgId":"","key":""}'
    #Tendering ='{"index":1,"size":20,"firstPageMenuId":"2018032700291334","purOrgStatus":"","purOrgCode":"","purType":"","orgId":"","key":""}'
    purchasedata='{"index":1,"size":40,"firstPageMenuId":"2018032900295987","purOrgStatus":"","purOrgCode":"","purType":"","orgId":"","key":""}'
    res = requests.post(url=url,data=purchasedata)
    result=json.loads(res.text)
    flag=0
    for i in range(20,0):
        listid=result['resultValue']["noteList"][i]['id']
        doctype=result['resultValue']["noteList"][i]['doctype']#doci-bid采购公告 doci-change：变更
        if str(listid) in codelist:
            pass
        else: 
            if doctype=="doci-bid":
                flag=1
                codelist.append(str(listid))
                getNoticeBid(listid)
            elif doctype=="doci-change":
                getChangeBid(listid)
            with open(os.getcwd()+'\\Res\\list.txt', "a") as f:
                f.write(str(listid)+"\r")
    os.popen('copy '+os.getcwd()+'\\Res\\采购信息.xlsx '+os.getcwd()+'\\国网招采.xlsx' )
    return flag

def quit(signum, frame):
    sys.exit()    

if __name__ =="__main__":
    if not os.path.exists(os.getcwd()+'\\Res\\采购信息.xlsx'):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title="物资采购"
        wb .create_sheet("服务采购")
        wb.save(os.getcwd()+'\\Res\\采购信息.xlsx')
        print("create xlsx")
    if not os.path.exists(os.getcwd()+'\\Res\\list.txt'):
        with open(os.getcwd()+'\\Res\\list.txt', "w") as f:
            f.write("")
        print("create txt")
    else:
        with open(os.getcwd()+'\\Res\\list.txt', "r") as f:
            for line in f:
                codelist.append(line.strip('\n'))
    signal.signal(signal.SIGINT, quit)                                
    signal.signal(signal.SIGTERM, quit)
    while True:
        print(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        if purchase()==1:
            win32api.MessageBox(0, "有新的采购信息了！", "提醒",win32con.MB_OK)
        time.sleep(300)
